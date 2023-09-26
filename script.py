############## This script splits up the QC sheet info needed into two categories: Reagent info and Results info

import bs4 as bs
import pandas as pd
from functools import reduce
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import time
from datetime import datetime

### declare the function for parsing specific query to return a list of dicts
def parse_xml(query,soup):
    try:
        search = soup.find_all(query)
        # parse content into dicts
        dicts=[]
        for data in search:
            if data.Carrier:
                data.attrs.update(data.Carrier.attrs)
            dicts.append(data.attrs)   
        return dicts
    except Exception as e: 
        print("\n Error in parse_xml: ",e,"\n")
        raise 
    
def clean_date(date):
    """ returns datetime object from a creation/modificationdate. initially intended for getting onboard time by reagentKit's creationdate - TestOrders"""
    date=date.split('.')[0]
    date=date.replace('T'," ")
    date=datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    date=datetime.date(date) # remove time so its just day
    return date
def get_assayDate(soup):
    """get the date that the assay was ran"""
    data=parse_xml('TestOrders',soup)
    data=data[0]['LastModificationDateTime']
    return data
def getReagents(soup):
    try:
        #query xml for reagents
        dicts=parse_xml('ReagentContainer',soup)
        #assemble into dataframe
        dfRs=pd.DataFrame.from_dict(dicts)
        dfRs=dfRs[['ReagentName','SerialNumber','LotNumber','Expiration','CreationDateTime']] # for reagents
        dfRs=dfRs.drop_duplicates()
        dfRs=dfRs.replace('(-) C', '(-) Ctrl')
        return dfRs
    except Exception as e: 
        print("\n Error in getReagents: ",e,"\n")
        raise
def getReagentKit(soup,test):
    try:
        KitMaterialNumbers={"HBV":"9040820190_A",
                 "HCV":"9040765190_A",
                 "HIV":"9040803190_A", 
                 "CTNG":"9040501190_A",
                 "SARS": "9343733190_A",
                 "HPV":"7460155190_A",}
        #query xml for kit
        dicts=parse_xml('InventoryItemTracking',soup)

        #assemble into dataframe
        dfK=pd.DataFrame.from_dict(dicts)
        dfK=dfK[['SerialNumber','LotNumber','Expiration','MaterialNumber','CreationDateTime']] # for reagents
        dfK=dfK.drop_duplicates()
        dfK['ReagentName']="Reagent Kit"
        
        ##Check for reagent kit by matching material number. 
        #material number may chnge with software updates for kits (i.e. CTNG in september 2023)
        #if the expected KitMaterialNumbers is not found:
            # it will auto-assign the kit from the first option, and therefore lot# needs to be double checked
            # it will also print the reagent table with lot#'s and materials #'s to confirm, and to update the new kit material #
        if len(dfK.loc[dfK['MaterialNumber']==KitMaterialNumbers[test]]) >0:
            dfK=dfK.loc[dfK['MaterialNumber']==KitMaterialNumbers[test]]
        else:
            print("The reagent kit's material number is not recognized.")
            dfK=dfK.dropna()
            print(dfK) #prints the reagent kit table to get the new kit # and confirm lot # to the physical box
            dfK=dfK.head(1) # get the first option by default which is often the reagent kit
            print("\nConfirm the reagent kit Lot# is: ", dfK['LotNumber'].item()," and update the reagent material number in the script based on the table above.")
        return dfK
    except Exception as e: 
        print("\n Error in getReagentsKit: ",e,"\n")
        raise
def rename_keys(my_dict,suffix):
    try:
        """ rename all keys in a dictionary with an added _suffix"""
        # create new key names from current by adding _suffix to it
        new_keys=[]
        suffix='_'+str(suffix)
        for x in my_dict:
            new=x+suffix
            new_keys.append(new)
        
        #create and return new dictionary
        d1 = dict( zip( list(my_dict.keys()), new_keys) )
        return {d1[oldK]: value for oldK, value in my_dict.items()}
    except Exception as e: 
        print("\n Error in rename_keys: ",e,"\n")
        raise
    
def infoFromTestOrder(one):
    try:
        """ get Specimen and TestResult data from a TestOrder tag"""
        row={}
        for x in one.children:
            # get sample info like barcode
            if x.name =="Specimen":
                for y in x.children:
                    row.update(y.attrs)

            # get test results info
            elif x.name =="TestResults":
                for num,y in enumerate(x.children):
                    if num==0:
                        row.update(y.attrs)
                    else:
                        new_dict= rename_keys(y.attrs,num)
                        row.update(new_dict)
                        multipleTargets=False
                   
        return row
    except Exception as e: 
        print("\n Error in infoFromTestOrder: ",e,"\n")
        raise
        
def getResults(soup):
    try:
        #outline general header terms desired
        GeneralHeaders=['CreationDateTime','Name','Barcode','FinalInterpretation','CT','Position','Info','SpecimenClass','Target','Value']
        #get testOrder info (each row of df)
        search = soup.find_all('TestOrder')
        for one in search:
            row = infoFromTestOrder(one)
            #convert row (dict) to df
            df=pd.DataFrame([row])
            #get all fields that contain a general header
            SpecificHeaders=[]
            for x in df:
                for y in GeneralHeaders:
                    if y in x:
                        SpecificHeaders.append(x)
            #filter df w/ desired fields
            df=df[SpecificHeaders]
            #try to append to main df if exists, except: create main_df
            try:
                dfResults=pd.concat([dfResults,df],axis=0)
            except:
                dfResults=df
        return dfResults
    except Exception as e: 
        print("\n Error in getResults: ",e,"\n")
        raise
        
def getTestName(dfResults):
    try:
        testName=""
        for x in dfResults:
            if "name" in x.lower():
                testName=dfResults[x].unique()[0]
        return testName
    except Exception as e: 
        print("\n Error in getTestName: ",e,"\n")
        raise
        
def addControlLabels(testName):
    try:
        ### add barcode values for control     
        # add barcode values for controls
        PosControlName=""
        HxV=['hiv','hbv','hcv']
        testName=testName.lower()
        if testName in HxV:
            print("control is HxV!!!")
            PosControlName="HxV L (+) C"
        elif "hpv" in testName:
            print("control is HPV!!!")
            PosControlName="HPV (+) C"
        elif "ct" in testName or "ng" in testName:
            print("control is CT/NG!!!")
            PosControlName="CT/NG (+) C"
        elif "tgt" in testName.lower():
            print("control is SARS!!!")
            PosControlName="SARS-CoV-2 (+) C"
        else:
            print("test name not found for labelling control samples")
        return PosControlName
    except Exception as e: 
        print("\n Error in addControlLabels: ",e,"\n")
        raise
        
def assignReagentVariables(dfRs,PosControlName):
    try:
        ### reagent info
        #kit lot, expiration, onboard time
        ReagentKitLot=str(dfRs.loc[dfRs['ReagentName']=='Reagent Kit']['LotNumber'].values[0])
        ReagentExpiration=str(dfRs.loc[dfRs['ReagentName']=='Reagent Kit']['Expiration'].values[0]).split('T')[0]
        #onboard time from creation date of kit and Last modification date of TestOrders
        ReagentOnboard=dfRs.loc[dfRs['ReagentName']=='Reagent Kit']['CreationDateTime'].values[0]
        #PC lot #
        PCLotNo=str(dfRs.loc[dfRs['ReagentName']==PosControlName]['LotNumber'].values[0])
        #NC lot#
        NCLotNo=str(dfRs.loc[dfRs['ReagentName']=='(-) Ctrl']['LotNumber'].values[0])
        return(ReagentKitLot,PCLotNo,NCLotNo,ReagentExpiration,ReagentOnboard)
    except Exception as e: 
        print("\n Error in assignReagentVariables: ",e,"\n")
        raise

def assignResultsVariablesHIV(dfResults):
    try:
        ### HIV
        #HPC CT
        HPC_CT=float(dfResults.loc[dfResults['Info']=='HxV H (+) C']['CT'].values[0])
        #HPC IU
        HPC_IU=float(dfResults.loc[dfResults['Info']=='HxV H (+) C']['Value'].values[0])
        #LPC CT
        LPC_CT=float(dfResults.loc[dfResults['Info']=='HxV L (+) C']['CT'].values[0])
        #LPC IU
        LPC_IU=float(dfResults.loc[dfResults['Info']=='HxV L (+) C']['Value'].values[0])
        #NC Result
        NC_CT=float(dfResults.loc[dfResults['Info']=='(-) C']['CT'].values[0])
        if NC_CT >0:
            pass
        else:
            NC_CT="None"
        return(HPC_CT,HPC_IU,LPC_CT,LPC_IU,NC_CT)
    except Exception as e: 
        print("\n Error in assignResultsVariablesHIV: ",e,"\n")
        raise
        
def assignResultsVariablesSARS(dfResults):
    try:
        tgt1_positives=0
        tgt2_positives=0
        invalidSamples=0
        for x in dfResults:
            if x== "FinalInterpretation":
                for y in dfResults[x].unique():
                    if "Positive"in y:
                        tgt1_positives=len(dfResults[x].loc[dfResults[x]==y])
            elif x== "FinalInterpretation_1":
                for y in dfResults[x].unique():
                    if "Positive"in y:
                        tgt2_positives=len(dfResults[x].loc[dfResults[x]==y])
                    elif "invalid"in y.lower():
                        invalidSamples=len(dfResults[x].loc[dfResults[x]==y])
        return(tgt1_positives,tgt2_positives,invalidSamples)  
    except Exception as e: 
        print("\n Error in assignResultVariablesSARS: ",e,"\n")
        raise
    
def assignResultsVariablesHPV(dfResults):
    try:
        ### HPV
        HPV16Positives=0
        HPV18Positives=0
        HPVotherPositives=0
        invalidSamples=0
        for x in dfResults:
            if "FinalInterpretation" in x:
                for y in dfResults[x].unique():
                    if "Positive"in y:
                        if "16" in y:
                            HPV16Positives=len(dfResults[x].loc[dfResults[x]==y]) # no minus 1 for controls, they are valid not positive
                        elif "18" in y:
                            HPV18Positives=len(dfResults[x].loc[dfResults[x]==y])
                        elif "other" in y.lower():
                            HPVotherPositives=len(dfResults[x].loc[dfResults[x]==y])
                    if "invalid"in y.lower():
                        invalidSamples=len(dfResults[x].loc[dfResults[x]==y])
        return(HPV16Positives,HPV18Positives,HPVotherPositives,invalidSamples)
    except Exception as e: 
        print("\n Error in assignResultsVariablesHPV: ",e,"\n")
        raise
        
def assignResultVariablesCTNG(dfResults):
    try:
        ### CTNG    
        swabSampleCount=0
        urineSampleCount=0
        thinprepSampleCount=0
        CTpositives=0
        NGpostives=0
        invalidSamples=0
        for x in dfResults:
            if "FinalInterpretation" in x:
                for y in dfResults[x].unique():
                    if "Positive"in y:
                        if "CT" in y:
                            CTpositives=len(dfResults[x].loc[dfResults[x]==y]) #no minus 1 for controls, they are valid not positive
                        elif "NG" in y:
                            NGpostives=len(dfResults[x].loc[dfResults[x]==y])
                    if "invalid"in y.lower():
                        invalidSamples=len(dfResults[x].loc[dfResults[x]==y])
            if "Info" in x:
                for y in dfResults[x].unique():
                    ymod=str(y).lower()
                    if "swab"in ymod:
                        swabSampleCount=len(dfResults[x].loc[dfResults[x]==y])
                    elif "urine"in ymod:
                        urineSampleCount=len(dfResults[x].loc[dfResults[x]==y])
                    elif "preserv"in ymod:
                        thinprepSampleCount=len(dfResults[x].loc[dfResults[x]==y])
        return(swabSampleCount,urineSampleCount,thinprepSampleCount,CTpositives,NGpostives,invalidSamples)
    except Exception as e: 
        print("\n Error in assignResultVariablesCTNG: ",e,"\n")
        raise
        
def prepQCdata(test,soup,path):
    """ load the QC file and find the line with cobas batch ID that matched the XML batch number"""
    try:
        #convert to linux
        cpath=path.replace('\\','/')
        # get batch # for joining 
        dicts=parse_xml('OrderGroup',soup)
        dfb=pd.DataFrame.from_dict(dicts)
        BatchNum=max([int(x) for x in dfb['OrderId'].unique()])
        #get and read correct sheet
        sheets={"HCV":"HCVrunQC",
               "HIV":"HIVrunQC",
               "HBV":"HBVrunQC",
               "HPV":"HPVrunQC",
               "SARS": "SARSrunQC",
               "CTNG":"CTNGrunQC"}
        sheet=sheets[test.lower().upper()]
        dftest=pd.read_excel(cpath, sheet_name=sheet,header=2)
        dftest=dftest.loc[dftest['CONTROL BATCH #']==BatchNum]
        if len(dftest)>0:
            pass
        else:
            print("Control batch ID not found")
        return dftest,sheet,BatchNum
    except Exception as e: 
        print("\n Error in prepQCdata: ",e,"\n")
        raise
        
def writeToQCSheets(dftest,sheet,BatchNum,path):
    try:
        ### use pyxl to write to spreadsheet in order to handle dateTime object
        #get lastrow number to write to
        wb = load_workbook(path)
        ws = wb[sheet]
        try:
            dftestValues=[x for x in dataframe_to_rows(dftest, index=False, header=False)][0]
        except:
            print("No batch number found in sheet: ",sheet,"for control batch ID: ",BatchNum,"\n Check that the COBAS batch # was entered in the correct sheet prior to running the script.")
        #Find Row for desired batch number
        for row in ws.iter_cols(min_col=3,max_col=3):
            for RowNum,cell in enumerate(row):
                if (cell.value == BatchNum):
                    targetRow=RowNum+1
        #Write values to excel sheet            
        for num,x in enumerate(ws[targetRow]):
            x.value=dftestValues[num]
        wb.save(path)
        wb.close()
    except Exception as e: 
        print("\n Error in writeToQCSheets: ",e,"\n")
        raise
        
def writeToResultsSheet(dfResults, test, BatchNum):
    """Function to write to results sheet. Requires: dfResults, test, BatchNum"""
    path="M:\MP Molecular Pathology\\NJ_Mol_Virology\\NJ Routine\QC Sheets\COBAS 8800\\All_results\\" + test + "_all_results_cobas" + ".csv"
    try:
        #load results file into df
        df=pd.read_csv(path)
        #check if the current batch is already loaded into the results sheet
        if BatchNum in df['Batch'].unique():
            print("Found previous results for this batch. Results sheet won't be updated")
        else:
            ####write data to results sheet
            #add cobas batch #
            dfResults['Batch']=BatchNum
            # re-order columns
            dfResults=dfResults[[x for x in df]]
            #update results sheet
            dfResults.to_csv(path, index=False, mode='a', header=False,)
    except Exception as e:
        print("\n","Error writing to results sheet: ",e)
        raise
        
def Main(name,path):
    # reading file content
    file = open(name, "r")
    contents = file.read()
    # parsing
    soup = bs.BeautifulSoup(contents, 'xml')
    file.close()
    
    ### get test name from file ##################################
    # requires test name in file
    # USE DefinitionId ID OR MaterialNumber NOT SERIAL NUMBER    
    tests=['HBV','HCV','HIV','HPV','CTNG','SARS']
    test=""
    try:
        for x in tests:
            if x.lower() in name.lower():
                print("test is: ",x)
                test=x
    except:
        print("test not found in file name. options are: ", tests)
    
    ### Reagents info extraction################################################################
    #get regeagents
    dfRs=getReagents(soup)
    #get reagentKit
    dfKit=getReagentKit(soup,test)
    #join reagents w kits info
    dfRs=pd.concat([dfRs,dfKit])
    
    ### Results info extraction ################################################################
    dfResults=getResults(soup) 
        
    #get testName
    testName=getTestName(dfResults)
    testName=testName.strip().lower()
    #add control labels for results
    PosControlName=addControlLabels(testName)
    sampleNum=len(dfResults)
    
    #assign reagent variables that will be written to QC sheet
    ReagentKitLot,PCLotNo,NCLotNo,ReagentExpiration,ReagentOnboard=assignReagentVariables(dfRs,PosControlName)
    AssayDate=clean_date(get_assayDate(soup))
    ReagentKitCreationDate=clean_date(ReagentOnboard)
    ReagentOnboard=( AssayDate-ReagentKitCreationDate).days # get the difference in days from the assay run date and reagent creationdate (onboard)    
    
    #assign result variables that will be written to QC sheet
    HxV=['hiv','hbv','hcv']
    if testName in HxV:
        HPC_CT,HPC_IU,LPC_CT,LPC_IU,NC_CT=assignResultsVariablesHIV(dfResults)
        dftest,sheet,BatchNum=prepQCdata(test,soup,path)
    elif "hpv" in testName:
        HPV16Positives,HPV18Positives,HPVotherPositives,invalidSamples=assignResultsVariablesHPV(dfResults)
        dftest,sheet,BatchNum=prepQCdata(test,soup,path)
    elif "ct" in testName or "ng" in testName:
        swabSampleCount,urineSampleCount,thinprepSampleCount,CTpositives,NGpostives,invalidSamples=assignResultVariablesCTNG(dfResults)
        dftest,sheet,BatchNum=prepQCdata(test,soup,path)
    elif "tgt" in testName.lower(): # SARS
        tgt1_positives,tgt2_positives,invalidSamples=assignResultsVariablesSARS(dfResults) 
        dftest,sheet,BatchNum=prepQCdata(test,soup,path)
    else:
        print("testName not found: ", testName)
        

    #assign reagent values
    dftest['auto REAGENT KIT LOT']=ReagentKitLot
    dftest['auto POSITIVE CTRL KIT LOT#']=PCLotNo
    dftest['auto NEGATIVE CTRL LOT#']=NCLotNo
    dftest['REAGENT KIT expiration']=ReagentExpiration
    dftest['REAGENT KIT onboard days']=ReagentOnboard
    #assign result values
    dftest['Samples + controls']=sampleNum
    
    if testName in HxV:
        dftest['CT VALUE OF HIGH POS']=HPC_CT
        dftest['HIGH POS CTRL RESULT (IU/mL)']=HPC_IU
        dftest['HIGH POS CTRL Result (Log IU/mL)']=np.log10(HPC_IU)
        dftest['CT VALUE OF LOW POS']=LPC_CT
        dftest['LOW POS CONTROL RESULT (IU/mL)']=LPC_IU
        dftest['LOW POS CONTROL (Log IU/mL)']=np.log10(LPC_IU)
        dftest['NEGATIVE CTRL RESULT']=NC_CT
    elif "hpv" in testName:
        dftest['OTHER HR HPV POSITIVE']=HPVotherPositives
        dftest['HPV 16 POSITIVE']=HPV16Positives
        dftest['HPV 18 POSITIVE']=HPV18Positives
        dftest['INVALID']=invalidSamples
    elif "ct" in testName or "ng" in testName:
        dftest['CT POSITIVE']=CTpositives
        dftest['NG POSITIVE']=NGpostives
        dftest['INVALID']=invalidSamples
        dftest['SWABS']=swabSampleCount
        dftest['URINES']=urineSampleCount
        dftest['THINPREPS']=thinprepSampleCount
    elif "tgt" in testName.lower(): # SARS
        dftest['Target1 positive']=tgt1_positives
        dftest['Target2 positive']=tgt2_positives
        dftest['INVALID']=invalidSamples
    else:
        print("test name not found")
    
    ### write values to QCsheet 
    writeToQCSheets(dftest,sheet,BatchNum,path)
    ###write to Results sheet
    writeToResultsSheet(dfResults, test, BatchNum)
    
if __name__=="__main__":
    qcSheetPath=str('M:\MP Molecular Pathology\\NJ_Mol_Virology\\NJ Routine\QC Sheets\COBAS 8800\\8800  Daily QC 2023.xlsx')
    #list files
    newFileFolder="new XML files/"
    for x in os.listdir(newFileFolder):
        #if file starts with 'b'
        if x[0].lower()=='b':
            print("~~~~~~~Processing file: ",x)
            path=newFileFolder+x 
            try:
                Main(path,qcSheetPath)
                newPath="old XML files/"+x  
                os.rename(path,newPath)
                print("Complete!",x,"\n")
            except Exception as e: 
                print("\n",e," \n error found, see above text \n \n")
    print("\n \n Program complete. You may now exit or Window will close shortly")        
    time.sleep(120)
